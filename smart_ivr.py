import pandas as pd
import openpyxl
from datetime import datetime as dt, timedelta as td
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

class ExcelWorker:
    def read_excel_create_dfs(self):
        print("Reading Excel...")
        month1_povici = pd.read_excel('fajlovi/povici_januari.xlsx')
        month2_povici = pd.read_excel('fajlovi/povici_fevruari.xlsx')
        month3_povici = pd.read_excel('fajlovi/povici_mart.xlsx')

        month1_prechki = pd.read_excel('fajlovi/prechki_januari.xlsx')
        month2_prechki = pd.read_excel('fajlovi/prechki_fevruari.xlsx')
        month3_precki = pd.read_excel('fajlovi/prechki_mart.xlsx')


        lista_povici = [month1_povici, month2_povici, month3_povici]
        self.df_povici = pd.concat(lista_povici, ignore_index=True)

        lista_precki = [month1_prechki, month2_prechki, month3_precki]
        self.df_precki = pd.concat(lista_precki, ignore_index=True)

        self.df_precki["Тип на пречката"] = (
        self.df_precki["Тип на пречката"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)  # fix NBSP
        .str.strip()
        .str.lower()
    )
        self.df_precki["Status nalog"] = self.df_precki["Status nalog"].astype(str).str.strip().str.lower()

        self.df_precki = self.df_precki[
            (self.df_precki["Status nalog"] != "откажан") &
            (self.df_precki["Тип на пречката"] != "network facing") &
            (self.df_precki["Класификација"] != "WHOLESALE") &
            (self.df_precki["Класификација"] != "MOBILE  POSTPAID") &
            (self.df_precki["Класификација"] != "MOBILE PREPAID") &
            (self.df_precki["Категорија"] == "Физичко лице")
        ]   

        self.df_precki["Контакт"] = self.df_precki["Контакт"].astype(str).str.replace(r'\.0$', '', regex=True).str.lstrip('0')
        self.df_precki["Контакт"] = self.df_precki["Контакт"].apply(lambda x: "+389" + x)

        self.df_precki.to_excel('fajlovi/prechki_combined.xlsx', index=False)

        self.df_povici = self.df_povici[
            (self.df_povici["Direction"] == "Inbound")
        ]

        return self.df_povici, self.df_precki

    def remove_FTTH_ready(self):
        print("Removing FTTH Ready...")
        ffth_ready = pd.read_excel('fajlovi/FTTH_Ready.xlsx')
        ffth_ready['CINUMS'] = ffth_ready['CINUMS'].astype(str).str.strip()
        self.df_precki['LineID'] = self.df_precki['LineID'].astype(str).str.strip()
        self.df_precki = self.df_precki[~self.df_precki['LineID'].isin(ffth_ready['CINUMS'])]
    
    def mark_slabi_linii(self):
        print("Marking weak lines...")
        slabi_linii = pd.read_excel('fajlovi/Slabi_Linii.xlsx', header=8)
        slabi_linii["Row Labels"] = slabi_linii["Row Labels"].astype(str).str.strip()
        self.df_precki['LineID'] = self.df_precki['LineID'].astype(str).str.strip()
       #add a column at the end of the table with "Slabi Linii", if it's found in the slabi_linii, mark it as "Yes", if not, mark it as "No"
        self.df_precki['слаба линија'] = self.df_precki['LineID'].apply(lambda x: 'Да' if x in slabi_linii['Row Labels'].values else 'Не')

    def styling(self):
        print("Styling Excel...")
        column_widths = {
            'LineID': 15,
            'телефонски број': 30,
            'отворени пречки': 30,
            'број на повици во контакт центар': 40,
            'слаба линија': 20
        }

        workbook = load_workbook('fajlovi/prechki_povici_combined.xlsx')
        sheet = workbook.active
        # get headerr
        header = [cell.value for cell in sheet[1]]
        for col_name, width in column_widths.items():
            if col_name in header:
                col_idx = header.index(col_name) + 1  # 1-based
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = width
        workbook.save('fajlovi/prechki_povici_combined.xlsx')

        workbook = load_workbook('fajlovi/CINUMS.xlsx')
        sheet = workbook.active
        cinums_width = 20

        reason_code_width = 30

        header = [cell.value for cell in sheet[1]]
        if 'CINUMS' in header:
            col_idx = header.index('CINUMS') + 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = cinums_width
        if 'Reason Code' in header:
            col_idx = header.index('Reason Code')
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = reason_code_width

        workbook.save('fajlovi/CINUMS.xlsx')

    def create_report(self):
        print("Creating report...")
        counter_obj = Counter(self.df_povici, self.df_precki)

        counted_prechki_df = counter_obj.count_precki()
        counted_povici_df = counter_obj.count_povici()

        merged = (
            counted_prechki_df
            .merge(counted_povici_df, on='телефонски број', how='outer')
            .fillna(0)
            .astype({'отворени пречки': 'int', 'број на повици во контакт центар': 'int'})
            .query('`отворени пречки` > 0')
            .query('`број на повици во контакт центар` > 1')
        )
        
        # promena na redosled na koloni
        merged = merged[['LineID', 'телефонски број', 'отворени пречки', 'број на повици во контакт центар', 'слаба линија']]

        merged.to_excel('fajlovi/prechki_povici_combined.xlsx', index=False)
        final_df = pd.read_excel('fajlovi/prechki_povici_combined.xlsx')
        return final_df


class Counter:
    def __init__(self, df_povici, df_precki):
        self.df_povici = df_povici
        self.df_precki = df_precki

    def count_precki(self):
        print("Counting faults...")
        # normalize
        df_temp = self.df_precki.copy()
        df_temp['телефонски број'] = (
            df_temp['Контакт']
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.strip()
        )
        
        # group
        result = (
            df_temp
            .groupby('телефонски број')
            .agg({'LineID': 'first', 'Контакт': 'size', 'слаба линија': 'first'})
            .reset_index()
            .rename(columns={'Контакт': 'отворени пречки'})
        )
        
        result = result[['LineID', 'телефонски број', 'отворени пречки', 'слаба линија']]
        return result

    def count_povici(self):
        print("Counting calls...")
        column = 'TBP_ANI (Case) (Old Value)'
        if column not in self.df_povici.columns:
            raise KeyError(f"Required column '{column}' missing in df_povici")

        df = (
            self.df_povici[column]
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.strip()
            .value_counts()
            .reset_index()
        )
        df.columns = ['телефонски број', 'број на повици во контакт центар']
        return df


class Histogram:
    def __init__(self, final_df):
        self.final_df = final_df
        self.counter_obj = Counter(None, None)

    def plot_raw(self):
        plt.figure(figsize=(10, 6))
    
    def plot_average(self):
        average_calls = self.final_df['број на повици во контакт центар'].mean()
        plt.figure(figsize=(6, 4))

    def scatter_plot(self):
        plt.figure(figsize=(10, 6))
        plt.scatter(self.final_df['отворени пречки'], self.final_df['број на повици во контакт центар'])
        plt.xlabel('Отворени пречки')
        plt.ylabel('Број на повици во контакт центар')
        plt.title('Отворени пречки vs Број на повици во контакт центар')
        plt.grid()
        plt.show()
        

def main():
    excel_obj = ExcelWorker()
    print("Starting analysis...")
    excel_obj.read_excel_create_dfs()
    print("Removing FTTH Ready entries...")
    excel_obj.remove_FTTH_ready()
    excel_obj.mark_slabi_linii()
    final_df = excel_obj.create_report()
    excel_obj.styling()
    
    filtered_df = final_df[(final_df['отворени пречки'] >= 2) & (final_df['слаба линија'] != 'Да')]
    cinums_df = pd.DataFrame({
        'CINUMS': filtered_df['LineID'],
        'Reason Code': 'RepeatedTCCAgent'
    })
    cinums_df.to_excel('fajlovi/CINUMS.xlsx', index=False)
    
    histogram_obj = Histogram(final_df)
    print("Plotting Diagrams...")
    histogram_obj.scatter_plot()

main()