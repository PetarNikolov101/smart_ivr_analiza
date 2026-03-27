import pandas as pd
import openpyxl
from datetime import datetime as dt, timedelta as td
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

def read_excel_create_dfs():
    month1_povici = pd.read_excel('fajlovi/povici_januari.xlsx')
    month2_povici = pd.read_excel('fajlovi/povici_fevruari.xlsx')
    # month3_povici = pd.read_excel('fajlovi/povici_mart.xlsx')

    month1_prechki = pd.read_excel('fajlovi/prechki_januari.xlsx')
    month2_prechki = pd.read_excel('fajlovi/prechki_fevruari.xlsx')
    # month3_precki = pd.read_excel('fajlovi/preciki_mart.xlsx')

    lista_povici = [month1_povici, month2_povici]
    df_povici = pd.concat(lista_povici, ignore_index=True)

    lista_precki = [month1_prechki, month2_prechki]
    df_precki = pd.concat(lista_precki, ignore_index=True)

    df_precki["Тип на пречката"] = (
    df_precki["Тип на пречката"]
    .astype(str)
    .str.replace("\xa0", " ", regex=False)  # fix NBSP
    .str.strip()
    .str.lower()
)
    df_precki["Status nalog"] = df_precki["Status nalog"].astype(str).str.strip().str.lower()

    df_precki = df_precki[
       # (df_precki["Satus nalog"] != "откажан") &
        (df_precki["Тип на пречката"] != "network facing") &
        (df_precki["Класификација"] != "WHOLESALE") &
        (df_precki["Класификација"] != "MOBILE  POSTPAID") &
        (df_precki["Класификација"] != "MOBILE PREPAID") &
        (df_precki["Категорија"] == "Физичко лице")
    ]   

    df_precki["Контакт"] = df_precki["Контакт"].astype(str).str.replace(r'\.0$', '', regex=True).str.lstrip('0')
    df_precki["Контакт"] = df_precki["Контакт"].apply(lambda x: "+389" + x)

    df_precki.to_excel('fajlovi/prechki_combined.xlsx', index=False)

    return df_povici, df_precki

class Counter:
    def __init__(self, df_povici, df_precki):
        self.df_povici = df_povici
        self.df_precki = df_precki

    def count_precki(self):
        # Normalize phone numbers
        df_temp = self.df_precki.copy()
        df_temp['телефонски број'] = (
            df_temp['Контакт']
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.strip()
        )
        
        # Group by normalized phone: count & get first LineID
        result = (
            df_temp
            .groupby('телефонски број')
            .agg({'LineID': 'first', 'Контакт': 'size'})
            .reset_index()
            .rename(columns={'Контакт': 'отворени пречки'})
        )
        
        result = result[['LineID', 'телефонски број', 'отворени пречки']]
        return result

    def count_povici(self):
        column = 'TBP_ANI (Case) (Old Value)'
        if column not in self.df_povici.columns:
            raise KeyError(f"Required column '{column}' missing in df_povici")

        df = (
            self.df_povici[column]
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.strip()
            .value_counts()
            .reset_index()
        )
        df.columns = ['телефонски број', 'број на повици во контакт центар']
        return df


class Histogram:
    def __init__(self, final_df):
        self.final_df = final_df
        self.counter_obj = Counter(None, None)

    def plot_raw(self):
        plt.figure(figsize=(10, 6))
    
    def plot_average(self):
        average_calls = self.final_df['број на повици во контакт центар'].mean()
        plt.figure(figsize=(6, 4))

    def scatter_plot(self):
        plt.figure(figsize=(10, 6))
        plt.scatter(self.final_df['отворени пречки'], self.final_df['број на повици во контакт центар'])
        plt.xlabel('Отворени пречки')
        plt.ylabel('Број на повици во контакт центар')
        plt.title('Отворени пречки vs Број на повици во контакт центар')
        plt.grid()
        plt.show()
        

def styling():
    column_widths = {
        'LineID': 15,
        'телефонски број': 30,
        'отворени пречки': 30,
        'број на повици во контакт центар': 40
    }
    workbook = load_workbook('fajlovi/prechki_povici_combined.xlsx')
    sheet = workbook.active
    # Get header row (assuming row 1)
    header = [cell.value for cell in sheet[1]]
    for col_name, width in column_widths.items():
        if col_name in header:
            col_idx = header.index(col_name) + 1  # 1-based
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width
    workbook.save('fajlovi/prechki_povici_combined.xlsx')

def create_report(df_povici, df_precki):
    counter_obj = Counter(df_povici, df_precki)

    counted_prechki_df = counter_obj.count_precki()
    counted_povici_df = counter_obj.count_povici()

    merged = (
        counted_prechki_df
        .merge(counted_povici_df, on='телефонски број', how='outer')
        .fillna(0)
        .astype({'отворени пречки': 'int', 'број на повици во контакт центар': 'int'})
        .query('`отворени пречки` > 0')
        .query('`број на повици во контакт центар` > 1')
    )
    
    # Reorder columns: LineID first, then телефонски број, then counts
    merged = merged[['LineID', 'телефонски број', 'отворени пречки', 'број на повици во контакт центар']]

    print('Combined counts (prechki + povici):')
    print(merged)
    merged.to_excel('fajlovi/prechki_povici_combined.xlsx', index=False)
    final_df = pd.read_excel('fajlovi/prechki_povici_combined.xlsx')
    return final_df

def main():
    df_povici, df_precki = read_excel_create_dfs()
    df_povici.info()
    df_precki.info()

    final_df = create_report(df_povici, df_precki)
    styling()
    
    # Create CINUMS file with IDs
    cinums_df = pd.DataFrame({'CINUMS': final_df['LineID']})
    cinums_df.to_excel('fajlovi/CINUMS.xlsx', index=False)
    
    histogram_obj = Histogram(final_df)
    histogram_obj.scatter_plot()

main()